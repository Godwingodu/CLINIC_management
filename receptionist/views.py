from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.contrib import messages

from receptionist .models import Invoice,Patient,Therapist,TherapistWorkingDay,TherapistLogin,InvoiceItem,ExportedInvoice,ExportFormat,TwilioRestException,Appointment,APPOINTMENT_STATUS,Speciality,Branch
from .constants import STATUS_CHOICES,PAYMENT_MODE_CHOICES

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
from .forms import *

# from .forms import AddPatientForm, PatientFilterForm
from django.views import View

# from .forms import AddPatientForm, PatientFilterForm
from django.db.models import Q
from django.db import transaction


from django.http import HttpResponseForbidden
from django.utils import timezone
from calendar import monthrange, monthcalendar
from collections import defaultdict
from datetime import date


logger = logging.getLogger(__name__)

import logging
import io



def receptionist_dashboard(request):
    return render(request, 'reception/receptionist_dashboard.html')

# Adding a New Patient
def add_new_patient(request):
    if request.method == 'POST':
        form = AddPatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.name} has been successfully added!')
            return redirect('patient_list')
    else:
        form = AddPatientForm()

    return render(request, 'reception/add_new_patient.html', {'form': form})

# Listing Patients with Filtering and Pagination
def patient_list(request):
    patients = Patient.objects.all().order_by('-date_added')
    form = PatientFilterForm(request.GET)
    
    if form.is_valid():
        patients = form.filter_queryset(patients)

    if 'export_excel' in request.GET:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="patients.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"
        
        # Excel headers
        headers = ["ID", "Name", "Address", "Date of Birth", "Phone", "Email", "Branch"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
        
        # Excel data rows
        for row_num, patient in enumerate(patients, 2):
            ws[f"A{row_num}"] = patient.patient_id
            ws[f"B{row_num}"] = patient.name
            ws[f"C{row_num}"] = patient.address
            ws[f"D{row_num}"] = patient.date_of_birth
            ws[f"E{row_num}"] = patient.phone
            ws[f"F{row_num}"] = patient.email
            ws[f"G{row_num}"] = patient.branch.name if patient.branch else ''
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.read())


    elif 'export_pdf' in request.GET:
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="patients.pdf"'
        
        buffer = BytesIO()
        p = SimpleDocTemplate(buffer, pagesize=letter)
        
        # PDF Table Setup
        data = [["ID", "Name", "Address", "Date of Birth", "Phone", "Email", "Branch"]]
        
        for patient in patients:
            data.append([
                patient.patient_id,
                patient.name,
                patient.address,
                patient.date_of_birth,
                patient.phone,
                patient.email,
                patient.branch.name if patient.branch else ''
            ])
        
        t = Table(data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elems = []
        elems.append(t)
        p.build(elems)
        
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        return response
        
        
    
    paginator = Paginator(patients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'reception/patient_list.html', {'form': form, 'page_obj': page_obj})

# Viewing a Single Patient Profile
class PatientProfileView(View):
    def get(self, request, pk):
        patient = get_object_or_404(Patient, id=pk)
        return render(request, 'reception/patient_profile.html', {'patient': patient})
    
def get_patient_details(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)
    data = {
        'id': patient.id,
        'name': patient.name,
        'address': patient.address,
        'phone': patient.phone,
        'date_of_birth': patient.date_of_birth,
        'gender': patient.gender,
        'height': str(patient.height),
        'weight': str(patient.weight),
        'blood_group': patient.blood_group,
        'notes': patient.notes,
        'email': patient.email,
        # add other fields
    }
    return JsonResponse(data)








# Updating Patient Information
def update_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)


    if request.method == 'POST':
        form = AddPatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, f'Patient {patient.name} has been successfully updated!')
            return redirect('patient_list')
    else:
        form = AddPatientForm(instance=patient)

    return render(request, 'reception/update_patient.html', {'form': form, 'patient': patient})

# Deleting a Patient
def delete_patient(request, patient_id):
    patient = get_object_or_404(Patient, id=patient_id)

    
    if request.method == 'POST':
        patient.delete()
        messages.success(request, f'Patient {patient.name} has been successfully deleted!')
        return redirect('patient_list')

    return render(request, 'reception/delete_patient.html', {'patient': patient})



#physiotherapist views 


from django.db import IntegrityError  # Import IntegrityError

@transaction.atomic  # Ensures database integrity
def add_new_therapist(request):
    context = {}
    if request.method == "POST":
        form = AddTherapistForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                new_therapist = form.save(commit=False)
                new_therapist.save()

                # Check for duplicate specialities before saving them
                existing_specialities = new_therapist.specialities.all().values_list('id', flat=True)
                new_specialities = form.cleaned_data['specialities']

                for speciality in new_specialities:
                    if speciality.id in existing_specialities:
                        raise IntegrityError(f"Duplicate speciality entry for therapist_id {new_therapist.id}")

                # Save many-to-many relationships
                # form.save_m2m()

                user = User.objects.create_user(
                    username=form.cleaned_data['email'],
                    email=form.cleaned_data['email'],
                    password=form.cleaned_data['password1'],
                )
                TherapistLogin.objects.create(
                    therapist=new_therapist,
                    user=user,
                    profile_photo=form.cleaned_data['profile_photo']
                )


                # Handle the working days and time slots for the therapist
                working_days_selected = form.cleaned_data['working_days']
                for day in working_days_selected:
                    # Read the time slots from hidden form field for each working day
                    time_slot_str = request.POST.get(f'time_slot_{day.day_name.lower()}')
                    
                    # If there are time slots for this day
                    if time_slot_str:
                        # Split the time slots into a list
                        time_slots = time_slot_str.split(", ")
                        
                        # Create a TherapistWorkingDay object for this therapist and day
                        working_day_rel = TherapistWorkingDay.objects.create(
                            therapist=new_therapist, working_day=day)
                        
                        # Create TimeSlot objects and link them to the working day
                        for time in time_slots:
                            start, end = time.split("-")
                            new_time_slot = TimeSlot.objects.create(start_time=start, end_time=end)
                            working_day_rel.time_slots.add(new_time_slot)
                    
                    print("Therapist added successfully")
                    return redirect('therapist_list')

            except IntegrityError as e:  # Changed Exception to IntegrityError
                context['error'] = f"Integrity Error: {e}"
                transaction.rollback()
                print(f"Error: {e}")
            except Exception as e:  # General exception
                context['error'] = str(e)
                transaction.rollback()
    else:
        form = AddTherapistForm()

    context['form'] = form
    context['time_slot_fields'] = form.time_slot_fields  # Add the new attribute to the context
    return render(request, 'reception/add_new_therapist.html', context)



def add_new_speciality(request):
    # Initialize an empty form and retrieve all existing specialities
    form = AddSpecialityForm()
    specialities = Speciality.objects.all()

    # Check if a POST request has been made
    if request.method == 'POST':
        # Determine the action based on a hidden field in the form
        action = request.POST.get('action')
        
        # Handle 'add' action to add a new speciality
        if action == 'add':
            form = AddSpecialityForm(request.POST)
            if form.is_valid():
                new_speciality = form.save()
                messages.success(request, f'Successfully added {new_speciality.name}.')
                return redirect('add_new_speciality')  # Redirect to clear form
        
        # Handle 'update' action to update an existing speciality
        elif action == 'update':
            speciality_id = request.POST.get('speciality_id')
            speciality = get_object_or_404(Speciality, id=speciality_id)
            form = AddSpecialityForm(request.POST, instance=speciality)
            if form.is_valid():
                form.save()
                messages.success(request, f'Successfully updated {speciality.name}.')
                return redirect('add_new_speciality')  # Redirect to show updated list

        # Handle 'delete' action to delete an existing speciality
        elif action == 'delete':
            speciality_id = request.POST.get('speciality_id')
            Speciality.objects.filter(id=speciality_id).delete()
            messages.success(request, 'Successfully deleted speciality.')
            return redirect('add_new_speciality')  # Redirect to show updated list

    # Render the template with the form and existing specialities
    return render(request, 'reception/add_new_speciality.html', {'form': form, 'specialities': specialities})




def list_therapists(request):
    # Receive query parameters for filtering and pagination
    search_query = request.GET.get('search', '')
    is_active_filter = request.GET.get('is_active', None)
    page_number = request.GET.get('page', 1)
    
    # Create the initial QuerySet
    queryset = Therapist.objects.select_related('branch').prefetch_related('specialities')
    
    # Apply search and filter conditions
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) | 
            Q(branch__name__icontains=search_query) |
            Q(specialities__name__icontains=search_query)
        ).distinct()
    
    if is_active_filter is not None:
        is_active = bool(int(is_active_filter))
        queryset = queryset.filter(is_active=is_active)
        
    # Apply pagination
    paginator = Paginator(queryset, 10)  # Show 10 therapists per page
    therapists = paginator.get_page(page_number)
    
    context = {
        'therapists': therapists,
        'search_query': search_query,
        'is_active_filter': is_active_filter
    }
    
    return render(request, 'reception/therapist_list.html', context)


def get_therapist_profile(request, therapist_id):
    try:
        therapist = Therapist.objects.select_related('branch').prefetch_related('specialities').get(id=therapist_id)
        specialities = [speciality.name for speciality in therapist.specialities.all()]
        profile_data = {
            'name': therapist.name,
            'branch': therapist.branch.name,
            'qualification': therapist.qualification,
            'specialities': specialities,
            'is_active': therapist.is_active,
            # Add other fields as needed
        }
        return JsonResponse(profile_data)
    except Therapist.DoesNotExist:
        return JsonResponse({'error': 'Therapist not found'}, status=404)




def add_prescription(request, appointment_id):
    appointment = get_object_or_404(Appointment, id=appointment_id)
    
    if request.method == 'POST':
        form = PrescriptionForm(request.POST, initial={
            'appointment': appointment,
        })
        if form.is_valid():
            prescription = form.save(commit=False)
            prescription.appointment = appointment
            prescription.therapist = appointment.therapist
            prescription.patient = appointment.patient
            prescription.save()
            messages.success(request, f'Prescription for {prescription.patient} successfully added.')
            return redirect('appointment_detail', appointment_id=appointment.id)
        else:
            messages.error(request, 'There was an error in your form.')
    else:
        form = PrescriptionForm(initial={
            'appointment': appointment,
            'therapist': appointment.therapist,
            'patient': appointment.patient,
        })
    
    return render(request, 'reception/add_prescription.html', {'form': form})




def get_appointment_detail(request, appointment_id):
    try:
        appointment = Appointment.objects.get(id=appointment_id)
        appointment_detail = {
            'therapist': appointment.therapist.name,
            'patient': appointment.patient.name,
        }
        return JsonResponse(appointment_detail)
    except Appointment.DoesNotExist:
        return JsonResponse({'error': 'Appointment not found'}, status=404)



#appointment views

# Import necessary modules

import json











# View function to schedule an appointment
def schedule_appointment(request, patient_id=None):
    # Check if there's a patient_id passed in the URL
    if patient_id:
        patient = get_object_or_404(Patient, id=patient_id)
        form = AppointmentForm(initial={'patient': patient})
    else:
        form = AppointmentForm()

    # Handle AJAX requests for available slots
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # Replaced request.is_ajax()
        therapist_id = request.GET.get('therapist_id')
        appointment_date = request.GET.get('appointment_date')
        if therapist_id and appointment_date:
            return get_available_slots(request, therapist_id, appointment_date)
    
    # Handle POST data when the form is submitted
    if request.method == "POST":
        form = AppointmentForm(request.POST)
        if form.is_valid():
            new_appointment = form.save(commit=False)
            new_appointment.status = 'Pending'
            new_appointment.save()
            
            messages.success(request, 'Appointment is pending approval.')
            return redirect('receptionist_dashboard')
        else:
            messages.error(request, 'An error occurred. Please try again.')
    
    therapists = Therapist.objects.all()
    therapists_list = []
    for therapist in therapists:
        therapist_dict = {
            'id': therapist.id,
            'name': therapist.name,
            'working_days': list(therapist.working_days.values_list('weekday_number', flat=True)),
        }
        therapists_list.append(therapist_dict)
                
    context = {
        'form': form,
        'therapists': therapists_list,
    }
      
    return render(request, 'reception/schedule_appointment.html', context)

# Function to get available slots
def get_available_slots(request, therapist_id, appointment_date):
    therapist = get_object_or_404(Therapist, id=therapist_id)
    existing_appointments = Appointment.objects.filter(
        therapist=therapist,
        appointment_date=appointment_date,
        status__in=['Pending', 'Accepted']
    )
    booked_slots = [appointment.time_slot.id for appointment in existing_appointments]
    
    # Fetch all available slots for this therapist
    all_slots = list(therapist.time_slots.values('id', 'name'))  # name is a field in time_slot model
    available_slots = []
    
    for slot in all_slots:
        slot_id = slot['id']
        is_booked = slot_id in booked_slots
        available_slots.append({
            'id': slot_id,
            'name': slot['name'],  
            'is_booked': is_booked
        })
    
    # Return the list of available slots as JSON
    return JsonResponse({'available_slots': available_slots})


def list_appointments(request):
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'GET' and is_ajax:
        search_query = request.GET.get('search', '')
        status_filter = request.GET.get('status', 'all')

        appointments = Appointment.objects.filter(
            Q(patient__user__first_name__icontains=search_query) |
            Q(therapist__name__icontains=search_query) |
            Q(appointment_date__icontains=search_query)
        )

        # Status Filtering for AJAX
        if status_filter == 'today':
            appointments = appointments.filter(appointment_date=timezone.now().date())
        elif status_filter == 'upcoming':
            appointments = appointments.filter(appointment_date__gt=timezone.now().date())
        elif status_filter == 'cancelled':
            appointments = appointments.filter(status='Cancelled')
        elif status_filter == 'completed':
            appointments = appointments.filter(status='Completed')

        appointments = appointments.values('id', 'patient__name', 'therapist__name', 'appointment_date', 'status', 'time_slot__name', 'patient__phone')

        return JsonResponse(list(appointments), safe=False)

    else:
        search_query = request.GET.get('search', '')
        status_filter = request.GET.get('status', 'all')
        selected_appointments = request.GET.get('selected_appointments', '')

        appointments = Appointment.objects.filter(
                Q(patient__name__icontains=search_query) |  # changed from patient__user__first_name to patient__name
                Q(therapist__name__icontains=search_query) |
                Q(appointment_date__icontains=search_query)
            )


        # Standard Status Filtering
        if status_filter == 'today':
            appointments = appointments.filter(appointment_date=timezone.now().date())
        elif status_filter == 'upcoming':
            appointments = appointments.filter(appointment_date__gt=timezone.now().date())
        elif status_filter == 'cancelled':
            appointments = appointments.filter(status='Cancelled')
        elif status_filter == 'completed':
            appointments = appointments.filter(status='Completed')

        paginator = Paginator(appointments, 10)
        page_number = request.GET.get('page')
        page_appointments = paginator.get_page(page_number)

        context = {
            'appointments': page_appointments,
            'search_query': search_query,
            'status_filter': status_filter
        }

        return render(request, 'reception/list_appointments.html', context)
    

# View function to update an appointment
def update_appointment(request, appointment_id):
    appointment = get_object_or_404(Appointment, id=appointment_id)
    form = AppointmentForm(instance=appointment)

    if request.method == "POST":
        form = AppointmentForm(request.POST, instance=appointment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Appointment updated successfully.')
            return redirect('list_appointments')
        else:
            messages.error(request, 'An error occurred. Please try again.')

    therapists = Therapist.objects.all()
    therapists_list = []
    for therapist in therapists:
        therapist_dict = {
            'id': therapist.id,
            'name': therapist.name,
            'working_days': list(therapist.working_days.values_list('weekday_number', flat=True)),
        }
        therapists_list.append(therapist_dict)

    context = {
        'form': form,
        'therapists': json.dumps(therapists_list),
        'appointment_id': appointment_id,
    }

    return render(request, 'reception/update_appointment.html', context)



# Function to delete an appointment
def delete_appointment(request, appointment_id):
    # Check if the request method is POST
    if request.method == "POST":
        try:
            # Fetch the appointment object to be deleted
            appointment = get_object_or_404(Appointment, id=appointment_id)

            # You can add authorization checks here. For example:
            # if request.user != appointment.patient.user and not request.user.is_staff:
            #     return HttpResponseForbidden("You don't have permission to delete this appointment.")
            
            appointment.delete()
            
            # Flash a success message
            messages.success(request, 'Appointment deleted successfully.')
        except Exception as e:
            # Flash an error message
            messages.error(request, f"An error occurred: {e}")
        return redirect('list_appointments')  # Change to your listing view name
    else:
        return HttpResponseForbidden("Invalid method.")


def calendar_view(request):
    selected_month = request.GET.get('month', timezone.now().month)
    selected_year = request.GET.get('year', timezone.now().year)

    try:
        selected_month = int(selected_month)
        selected_year = int(selected_year)
    except ValueError:
        return redirect('calendar_view')

    if selected_month < 1 or selected_month > 12 or selected_year < 1900 or selected_year > 3000:
        return redirect('calendar_view')

    first_day = date(selected_year, selected_month, 1)
    last_day_of_month = monthrange(selected_year, selected_month)[1]
    last_day = date(selected_year, selected_month, last_day_of_month)

    appointments_this_month = Appointment.objects.filter(appointment_date__range=(first_day, last_day))

    appointment_count_by_day = defaultdict(int)
    for appointment in appointments_this_month:
        appointment_count_by_day[appointment.appointment_date.day] += 1

    today = timezone.now().date()
    appointments_today = Appointment.objects.filter(appointment_date=today)

    calendar_data = monthcalendar(selected_year, selected_month)

    context = {
        'appointment_count_by_day': appointment_count_by_day,
        'appointments_today': appointments_today,
        'today': today,
        'calendar_data': calendar_data,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }

    return render(request, 'reception/calendar_view.html', context)


#ACCOUNTS VIEWS













# Optional parameter for appointment ID
@transaction.atomic
def create_invoice(request, appointment_id=None):
    # Check if an appointment_id is passed and if it exists
    if appointment_id:
        appointment = get_object_or_404(Appointment, id=appointment_id)
    else:
        appointment = None

    if request.method == "POST":
        form = InvoiceForm(request.POST)
        formset = InvoiceItemFormset(request.POST, prefix='items')

        if form.is_valid() and formset.is_valid():
            try:
                invoice = form.save(commit=False)
                if appointment:
                    invoice.appointment = appointment  # Set appointment if available
                invoice.save()  # This will also call recalculate_totals
                
                formset.instance = invoice
                formset.save()
                
                # Recalculate totals and save again
                invoice.recalculate_totals()
                invoice.save()

                logger.info(f"Invoice {invoice.invoice_number} created successfully.")
                messages.success(request, 'Invoice created successfully.')
                return redirect('invoice_list')  # Replace with the name of your invoice list view

            except Exception as e:
                logger.error(f"Failed to create invoice: {e}")
                messages.error(request, 'Failed to create invoice.')
                return redirect('create_invoice')  # Replace with the name of your invoice create view
        else:
            messages.warning(request, 'There are errors in the submitted form.')
    else:
        # Pre-fill the form with appointment data if available
        form = InvoiceForm(initial={'appointment': appointment}) if appointment else InvoiceForm()
        formset = InvoiceItemFormset(prefix='items')

    context = {
        'form': form,
        'formset': formset,
    }
    return render(request, 'reception/create_invoice.html', context)


def search_patients(request):
    query = request.GET.get('q', '')
    patients = Patient.objects.filter(name__icontains=query)[:5]  # Limit to 5 results
    patient_list = list(patients.values('name'))
    return JsonResponse({'patients': patient_list})



def invoice_list(request):
    search_form = SearchInvoiceForm(request.GET)
    query = request.GET.get('query', '')
    
    invoices_qs = Invoice.objects.all().order_by('-invoice_date')
    
    if query:
        invoices_qs = invoices_qs.filter(
            Q(invoice_number__icontains=query) |
            Q(patient__patient_id__icontains=query) |
            Q(therapist__name__icontains=query) |
            Q(appointment__appointment_date__icontains=query)
        )
    
    paginator = Paginator(invoices_qs, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    export_form = ExportFormatForm()

    context = {
        'search_form': search_form,
        'page_obj': page_obj,
        'export_form': export_form,
    }

    return render(request, 'reception/invoice_list.html', context)



def update_invoice_status(request, invoice_id):
    if request.is_ajax():  # Check if it's an AJAX request
        try:
            invoice = Invoice.objects.get(id=invoice_id)
        except Invoice.DoesNotExist:
            return JsonResponse({"error": "Invoice not found"}, status=404)

        new_status = request.POST.get('status')
        valid_statuses = [choice[0].lower() for choice in STATUS_CHOICES]
        if new_status.lower() in valid_statuses:
            with transaction.atomic():
                invoice.status = new_status
                invoice.save()
            return JsonResponse({"message": "Status updated successfully"})
        else:
            return JsonResponse({"error": "Invalid status"}, status=400)
    
    return JsonResponse({"error": "Invalid request"}, status=400)



def export_invoices(request):
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':  # Check if it's an AJAX request
        export_form = ExportFormatForm(request.POST)
        
        if export_form.is_valid():
            export_format = export_form.cleaned_data['export_format']
            selected_invoices = request.POST.getlist('selected_invoices')
            
            # Error handling for non-existent invoices
            missing_invoices = [inv for inv in selected_invoices if not Invoice.objects.filter(id=inv).exists()]
            if missing_invoices:
                return JsonResponse({"error": f"Could not find invoices with IDs: {', '.join(missing_invoices)}"}, status=404)

            if export_format.format_name == 'PDF':
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="invoices.pdf"'
                
                # Create the PDF object
                pdf = SimpleDocTemplate(
                    response,
                    pagesize=letter
                )

                # Create the table
                data = [['Invoice No', 'Patient ID', 'Name', 'Doctor', 'Appointment Date', 'Time', 'Status']]
                for invoice_id in selected_invoices:
                    invoice = Invoice.objects.get(id=invoice_id)
                    data.append([
                        invoice.invoice_number,
                        invoice.patient.patient_id,
                        invoice.patient.name,
                        invoice.therapist.name,
                        invoice.appointment.appointment_date,
                        invoice.appointment.time,  # Modify based on your actual model
                        invoice.status
                    ])

                # Add Table Style
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))

                # Add table to elements to build
                elements = []
                elements.append(table)

                # Generate PDF
                pdf.build(elements)

                return response
            
            return JsonResponse({"message": f"Invoices exported in {export_format.format_name} format."})
            
    return JsonResponse({"error": "Invalid request"}, status=400)


def paid_invoices(request):
    # Only filter invoices with 'paid' status
    invoices_qs = Invoice.objects.filter(status='paid').order_by('-invoice_date')
    
    # Paginator and other logic can remain similar to invoice_list()
    paginator = Paginator(invoices_qs, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'reception/paid_invoices.html', context)


def unpaid_invoices(request):
    # Only filter invoices with 'unpaid' status
    invoices_qs = Invoice.objects.filter(status='unpaid').order_by('-invoice_date')
    
    # Paginator and other logic can remain similar to invoice_list()
    paginator = Paginator(invoices_qs, 10)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
    }
    
    return render(request, 'reception/unpaid_invoices.html', context)


 # Make sure to import InvoiceItem

def view_paid_invoice(request, invoice_id):
    try:
        invoice = Invoice.objects.get(invoice_number=invoice_id, status='paid')  # Fetch the invoice by its number and status
        invoice.recalculate_totals()  # Update totals in case they are outdated
        invoice.save()  # Persist the changes
        items = InvoiceItem.objects.filter(invoice=invoice)  # Fetch the related items
    except Invoice.DoesNotExist:
        return redirect('paid_invoices')  # Redirect if the invoice does not exist or is not paid

    context = {
        'invoice': invoice,
        'items': items,
        'subtotal': invoice.subtotal,
        'tax': invoice.tax,
        'total': invoice.total_amount,
    }
    
    return render(request, 'reception/view_paid_invoice.html', context)




def view_unpaid_invoice(request, invoice_id):
    try:
        invoice = Invoice.objects.get(invoice_number=invoice_id, status='unpaid')  # Fetch the invoice by its number and status
        invoice.recalculate_totals()  # Update totals in case they are outdated
        invoice.save()  # Persist the changes
        items = InvoiceItem.objects.filter(invoice=invoice)  # Fetch the related items
    except Invoice.DoesNotExist:
        return redirect('unpaid_invoices')  # Redirect if the invoice does not exist or is not unpaid

    context = {
        'invoice': invoice,
        'items': items,
        'subtotal': invoice.subtotal,
        'tax': invoice.tax,
        'total': invoice.total_amount,
    }
    
    return render(request, 'reception/view_unpaid_invoice.html', context)






# reception physiotherapist views

def list_physiotherapists(request):
    # Receive query parameters for filtering and pagination
    search_query = request.GET.get('search', '')
    is_active_filter = request.GET.get('is_active', None)
    page_number = request.GET.get('page', 1)
    
    # Create the initial QuerySet
    queryset = Therapist.objects.select_related('branch').prefetch_related('specialities')
    
    # Apply search and filter conditions
    if search_query:
        queryset = queryset.filter(
            Q(name__icontains=search_query) | 
            Q(branch__name__icontains=search_query) |
            Q(specialities__name__icontains=search_query)
        ).distinct()
    
    if is_active_filter is not None:
        is_active = bool(int(is_active_filter))
        queryset = queryset.filter(is_active=is_active)
        
    # Apply pagination
    paginator = Paginator(queryset, 10)  # Show 10 therapists per page
    therapists = paginator.get_page(page_number)
    
    context = {
        'therapists': therapists,
        'search_query': search_query,
        'is_active_filter': is_active_filter
    }
    
    return render(request, 'reception/therapist_list.html', context)