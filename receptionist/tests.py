from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.views import View
from django.contrib import messages

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from io import BytesIO

from .models import Patient
from .forms import AddPatientForm, PatientFilterForm

# Adding a New Patient
def add_new_patient(request):
    if request.method == 'POST':
        form = AddPatientForm(request.POST, request.FILES)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.name} has been successfully added!')
            return redirect('patient_list')
    else:
        form = AddPatientForm()

    return render(request, 'add_new_patient.html', {'form': form})

# Listing Patients with Filtering and Pagination
def patient_list(request):
    patients = Patient.objects.all().order_by('-date_added')
    form = PatientFilterForm(request.GET)
    
    if form.is_valid():
        patients = form.filter_queryset(patients)

    if 'export_excel' in request.GET:
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="patients.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Patients"
        
        # Excel headers
        headers = ["ID", "Name", "Address", "Date of Birth", "Phone", "Email", "Branch"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
        
        # Excel data rows
        for row_num, patient in enumerate(patients, 2):
            ws[f"A{row_num}"] = patient.patient_id
            ws[f"B{row_num}"] = patient.name
            ws[f"C{row_num}"] = patient.address
            ws[f"D{row_num}"] = patient.date_of_birth
            ws[f"E{row_num}"] = patient.phone
            ws[f"F{row_num}"] = patient.email
            ws[f"G{row_num}"] = patient.branch.name if patient.branch else ''
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response.write(output.read())
        
        return response
    
    paginator = Paginator(patients, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'patient_list.html', {'form': form, 'page_obj': page_obj})

# Viewing a Single Patient Profile
class PatientProfileView(View):
    def get(self, request, patient_id):
        patient = get_object_or_404(Patient, patient_id=patient_id)
        return render(request, 'patient_profile.html', {'patient': patient})

# Updating Patient Information
def update_patient(request, patient_id):
    patient = get_object_or_404(Patient, patient_id=patient_id)

    if request.method == 'POST':
        form = AddPatientForm(request.POST, request.FILES, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, f'Patient {patient.name} has been successfully updated!')
            return redirect('patient_list')
    else:
        form = AddPatientForm(instance=patient)

    return render(request, 'update_patient.html', {'form': form, 'patient': patient})

# Deleting a Patient
def delete_patient(request, patient_id):
    patient = get_object_or_404(Patient, patient_id=patient_id)
    
    if request.method == 'POST':
        patient.delete()
        messages.success(request, f'Patient {patient.name} has been successfully deleted!')
        return redirect('patient_list')

    return render(request, 'delete_patient.html', {'patient': patient})
